from openpyxl import Workbook
import calendar
import CalendarStyle
from datetime import date

def SetMonthValues(workSheet, month, year=date.today().year):
  colIni = (month-1)*5+1
  days_month = calendar.monthrange(year, month)[1]
  maxRow = 5 + days_month
  for col in range(colIni,colIni+5):
    for row in range(5, maxRow+1):
      cell = workSheet.cell(row, col)

      if col==colIni+1:
        if row==5:
          cell.value = calendar.month_name[month]
        else:
          day_week = calendar.weekday(year, month, row-5)
          cell.value = calendar.day_name[day_week]
          
      elif col==colIni+2 and row > 5:
        cell.value = row-5
      elif col==colIni+3 and row == 5:
        cell.value = "HOURS"
  cell=workSheet.cell(maxRow+1,colIni+2,"TOTAL")

def SetFunctionMonth(workSheet, month, year=date.today().year):
  col = (month)*5-1
  days_month = calendar.monthrange(year, month)[1]
  maxRow = 5 + days_month

  InitialCell = ws.cell(6,col)
  EndCell = ws.cell(maxRow,col)

  workSheet.cell(maxRow+1,col,f"=SUM({InitialCell.coordinate}:{EndCell.coordinate})")

def SetYearValues(workSheet):
  for month in range(1,13):
    #Set Values
    SetMonthValues(workSheet, month)
    #Set Functions
    SetFunctionMonth(ws, month)

wb = Workbook()
ws = wb.active

#Set Values and functions
SetYearValues(ws)

#Set Styles
CalendarStyle.CellMergeStyle(ws)
CalendarStyle.YearBorderStyle(ws)

wb.save('Calendar.xlsx')