from openpyxl.styles import Border, Side, Alignment, Font

def MonthBorderStyle(workSheet, colIni=1, maxRow=37):

  mediumBorder = Side(style = "medium", color='00000000')
  doubleBorder = Side(style = "dashDot", color='00000000')
  bold = Font(name="Calibri",  color='FF0000', b=True)

  for col in range(colIni,colIni+5):
    for row in range(5, maxRow+1):
      cell = workSheet.cell(row, col)
      cell.font = Font(name="Calibri",  color='000000')
      
      if cell.value == "Sunday":
        cell.border = Border(bottom=doubleBorder, left=mediumBorder)
      
      if row==5:
        if col==colIni+1:
          cell.border = Border(top = mediumBorder, left=mediumBorder, bottom=mediumBorder)
          cell.alignment = Alignment(horizontal="center")
          cell.font=bold
        elif col==colIni+2:
          cell.border = Border(top = mediumBorder, right=mediumBorder, bottom=mediumBorder)
        elif col==colIni+3:
          cell.alignment = Alignment(horizontal="center")
          cell.border = Border(top = mediumBorder, left=mediumBorder, bottom=mediumBorder, right=mediumBorder)
      elif row==maxRow:
        if colIni < col < colIni+4:
          cell.border = Border(left=mediumBorder, right=mediumBorder, bottom=mediumBorder)
      else:
        if colIni < col < colIni+4:
          cell.border = Border(right=mediumBorder, left=mediumBorder)
      
  cell=workSheet.cell(maxRow+1,colIni+2)
  cell.font=bold

def CellMergeStyle(workSheet):
  for month in range(12):
        workSheet.merge_cells(start_row=5, end_row=5, start_column=5*month+2, end_column=5*month+3)

def YearBorderStyle(workSheet):
  DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
  for index, dayInMonth in enumerate(DAYS_IN_MONTH):
    MonthBorderStyle(workSheet, 5*index+1, 5+dayInMonth)
