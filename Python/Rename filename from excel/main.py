import os, sys
import openpyxl
import json

if getattr(sys, 'frozen', False):
    dpath = os.path.dirname(sys.executable)  # Path to the executable when frozen
else:
    dpath = os.path.dirname(__file__)  # Path to the script in development mode

files = os.listdir(dpath)
wb = None
for file in files:
    file_ext = os.path.splitext(file)[1].lower()  # Get the file extension
    if file_ext in ['.xlsx', '.xls']:
        print(f"Reading file: {file}")
        try:
            wb = openpyxl.load_workbook(os.path.join(dpath, file))
            print(f"Successfully loaded workbook: {file}")
            break  # Exit loop after loading the first valid Excel file
        except Exception as e:
            print(f"Error loading workbook: {e}")
            sys.exit(1)

if wb is None:
    print("No valid Excel file found in the directory.")
    sys.exit(1)

#os.chdir(dpath)

ws = wb.active
dic={}
for i in range(1,ws.max_row+1):
  cellKey = ws.cell(row=i, column=1).value
  cellValue = ws.cell(row=i, column=2).value
  dic[str(cellKey)]=str(cellValue)
keyList=list(dic.keys())

for key in keyList:
  old_file = os.path.join(dpath, key+'.jpeg')
  old_file2 = os.path.join(dpath, key+'.jpg')
  old_file3 = os.path.join(dpath, key+'.png')
  new_file = os.path.join(dpath, dic[key]+'.jpg')
  if os.path.isfile(old_file):
    os.replace(old_file, new_file)
  elif os.path.isfile(old_file2):
    os.replace(old_file2, new_file)
  elif os.path.isfile(old_file3):
    os.replace(old_file3, new_file)
  else:
    print(f"File {old_file} does not exist.")

#prints the dictionary 
print(json.dumps(dic,indent=4))
print(len(dic))
wb.close()


# Error

# about wkd of a executable
# https://stackoverflow.com/questions/19101470/why-is-the-working-directory-the-directory-of-the-executable-and-not-where-im-r
